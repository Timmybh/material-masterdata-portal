from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..auth import current_user, require_roles
from ..db import get_db
from ..models import Item, ItemGroup, MaterialRequest, MaterialType, Role, User
from ..schemas import CatalogIn, CatalogOut

router = APIRouter(prefix="/api/catalogs", tags=["catalogs"])
admin_only = require_roles(Role.ADMIN.value)
CATALOGS = {"material-types": MaterialType, "item-groups": ItemGroup}


def catalog_model(kind: str):
    model = CATALOGS.get(kind)
    if not model:
        raise HTTPException(404, "Danh mục không tồn tại")
    return model


def clean(payload: CatalogIn):
    return payload.code.strip().upper(), payload.name.strip()


def resolve_catalog_names(db: Session, material_type: str | None, item_group: str | None):
    if not material_type or not item_group:
        raise HTTPException(422, "Cần chọn Loại vật tư và Nhóm hàng")
    type_value = material_type.strip()
    group_value = item_group.strip()
    type_row = db.scalar(select(MaterialType).where((MaterialType.code == type_value.upper()) | (MaterialType.name == type_value)))
    group_row = db.scalar(select(ItemGroup).where((ItemGroup.code == group_value.upper()) | (ItemGroup.name == group_value)))
    if not type_row:
        raise HTTPException(422, "Loại vật tư không thuộc danh mục")
    if not group_row:
        raise HTTPException(422, "Nhóm hàng không thuộc danh mục")
    return type_row.name, group_row.name


@router.get("/{kind}", response_model=list[CatalogOut])
def list_catalog(kind: str, db: Session = Depends(get_db), _: User = Depends(current_user)):
    model = catalog_model(kind)
    return db.scalars(select(model).order_by(model.code)).all()


@router.post("/{kind}", response_model=CatalogOut, status_code=201)
def create_catalog(kind: str, payload: CatalogIn, db: Session = Depends(get_db), _: User = Depends(admin_only)):
    model = catalog_model(kind)
    code, name = clean(payload)
    if db.get(model, code):
        raise HTTPException(409, "Mã đã tồn tại")
    if db.scalar(select(model).where(model.name == name)):
        raise HTTPException(409, "Tên đã tồn tại")
    row = model(code=code, name=name)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/{kind}/{code}", response_model=CatalogOut)
def update_catalog(kind: str, code: str, payload: CatalogIn, db: Session = Depends(get_db), _: User = Depends(admin_only)):
    model = catalog_model(kind)
    old_code = code.strip().upper()
    row = db.get(model, old_code)
    if not row:
        raise HTTPException(404, "Không tìm thấy dữ liệu")
    new_code, name = clean(payload)
    if new_code != old_code and db.get(model, new_code):
        raise HTTPException(409, "Mã mới đã tồn tại")
    if db.scalar(select(model).where(model.name == name, model.code != old_code)):
        raise HTTPException(409, "Tên đã tồn tại")
    row.code = new_code
    row.name = name
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Không thể đổi mã đang được sử dụng")
    db.refresh(row)
    return row


@router.delete("/{kind}/{code}", status_code=204)
def delete_catalog(kind: str, code: str, db: Session = Depends(get_db), _: User = Depends(admin_only)):
    model = catalog_model(kind)
    row = db.get(model, code.strip().upper())
    if not row:
        raise HTTPException(404, "Không tìm thấy dữ liệu")
    values = {row.code, row.name}
    if kind == "material-types":
        used = db.scalar(select(MaterialRequest.id).where(MaterialRequest.item_type_name.in_(values)).limit(1)) or db.scalar(select(Item.id).where(Item.item_type_name.in_(values)).limit(1))
    else:
        used = db.scalar(select(MaterialRequest.id).where(MaterialRequest.item_group.in_(values)).limit(1)) or db.scalar(select(Item.id).where(Item.item_group_code.in_(values)).limit(1))
    if used:
        raise HTTPException(409, "Danh mục đang được sử dụng, không thể xóa")
    db.delete(row)
    db.commit()


def read_excel(file_bytes: bytes):
    try:
        ws = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True).active
    except Exception as exc:
        raise HTTPException(422, "File Excel không hợp lệ") from exc
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(422, "File Excel không có dữ liệu")
    normalized = {str(value or "").strip().lower(): index for index, value in enumerate(rows[0])}
    code_index = next((normalized[x] for x in ("mã", "ma", "code") if x in normalized), None)
    name_index = next((normalized[x] for x in ("tên", "ten", "name") if x in normalized), None)
    if code_index is None or name_index is None:
        raise HTTPException(422, "Excel phải có hai cột Mã và Tên")
    result = []
    for number, values in enumerate(rows[1:], start=2):
        code = str(values[code_index] or "").strip().upper()
        name = str(values[name_index] or "").strip()
        if not code and not name:
            continue
        if not code or not name:
            raise HTTPException(422, f"Dòng {number} thiếu Mã hoặc Tên")
        result.append((code, name))
    if not result:
        raise HTTPException(422, "File Excel không có dòng dữ liệu")
    return result


@router.post("/{kind}/import")
async def import_catalog(kind: str, file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(admin_only)):
    model = catalog_model(kind)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(422, "Chỉ hỗ trợ file .xlsx")
    rows = read_excel(await file.read())
    created = updated = 0
    for code, name in rows:
        row = db.get(model, code)
        if row:
            row.name = name
            updated += 1
        else:
            db.add(model(code=code, name=name))
            created += 1
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(409, "File có mã hoặc tên bị trùng") from exc
    return {"created": created, "updated": updated, "total": len(rows)}
