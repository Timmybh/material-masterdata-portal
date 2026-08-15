from io import BytesIO
from uuid import UUID
from fastapi import APIRouter,Depends,HTTPException,Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session,joinedload
from ..auth import require_roles
from ..db import get_db
from ..models import MaterialRequest,RequestStatus,Role,User
from ..notifications import notify
from ..schemas import CompleteIn,ReasonIn,RequestOut
from ..workflow import transition
router=APIRouter(prefix="/api/accounting",tags=["accounting"]);allowed=require_roles(Role.ACCOUNTING.value)
BRAVO_HEADERS=["Code","ParentId","IsGroup","Name","ItemTypeName","Name2","ItemCustom","IsCustoms","IsItemWithColor","IsItemWithSize","IsItemWithArt","IsItemWithProductCostId","IsItemWithBizDocId_C2","IsItemWithSymmetrical","IsItemWithColorProduct","ParentCode","ItemGroupCode","KindCode","CustomerCode","ProductCostInfo","ProductItemCode","BranchCode","Ghi chú"]

def bravo_value_row(r):
    purpose_note=(r.notes or r.purpose or "").strip() or None
    return [
        r.result_item_code,None,False,r.item_name,r.item_type_name,r.technical_specs,
        None,False,r.with_color,r.with_size,r.with_art,False,False,False,False,
        r.parent_code,r.item_group,r.kind_code,r.customer_code,None,None,r.branch_code,
        purpose_note,
    ]

def build_bravo_workbook(rows):
    wb=Workbook();ws=wb.active;ws.title="vB20Item";ws.append(BRAVO_HEADERS)
    for r in rows:ws.append(bravo_value_row(r))
    header_fill=PatternFill("solid",fgColor="D9D9D9")
    for cell in ws[1]:
        cell.fill=header_fill;cell.font=Font(bold=True);cell.alignment=Alignment(horizontal="center",vertical="center")
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    widths={"A":18,"D":36,"E":22,"F":42,"Q":20,"W":45}
    for column,width in widths.items():ws.column_dimensions[column].width=width
    for column in ("D","E","F","W"):
        for cell in ws[column]:cell.alignment=Alignment(vertical="top",wrap_text=True)
    return wb

def load(db,rid):
    req=db.scalar(select(MaterialRequest).options(joinedload(MaterialRequest.requester)).where(MaterialRequest.id==rid))
    if not req:raise HTTPException(404,"Không tìm thấy yêu cầu")
    return req
@router.get("/requests",response_model=list[RequestOut])
def queue(db:Session=Depends(get_db),_:User=Depends(allowed)):
    return db.scalars(select(MaterialRequest).options(joinedload(MaterialRequest.requester)).where(MaterialRequest.status==RequestStatus.MASTERDATA_APPROVED.value).order_by(MaterialRequest.submitted_at)).all()
@router.get("/export.xlsx")
def export_excel(ids:str=Query(min_length=1),db:Session=Depends(get_db),_:User=Depends(allowed)):
    try:selected_ids=[UUID(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:raise HTTPException(422,"Danh sách yêu cầu xuất không hợp lệ")
    if not selected_ids:raise HTTPException(422,"Vui lòng chọn ít nhất một yêu cầu để xuất Excel")
    q=select(MaterialRequest).options(joinedload(MaterialRequest.requester)).where(MaterialRequest.status==RequestStatus.MASTERDATA_APPROVED.value)
    q=q.where(MaterialRequest.id.in_(selected_ids))
    rows=db.scalars(q.order_by(MaterialRequest.submitted_at)).all();wb=build_bravo_workbook(rows);bio=BytesIO();wb.save(bio);bio.seek(0)
    return StreamingResponse(bio,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=BRAVO_vB20Item.xlsx"})
@router.post("/requests/{rid}/complete",response_model=RequestOut)
def complete(rid:UUID,payload:CompleteIn,db:Session=Depends(get_db),user:User=Depends(allowed)):
    req=load(db,rid)
    if req.status!=RequestStatus.MASTERDATA_APPROVED.value:raise HTTPException(409,"Yêu cầu chưa ở trạng thái Kế toán xử lý")
    req.result_item_code=payload.item_code.strip();req.accounting_note=(payload.note or "").strip() or None
    audit_note=f"Mã mới: {req.result_item_code}"+(f" | Ghi chú: {req.accounting_note}" if req.accounting_note else "")
    req=transition(db,req,user,"ACCOUNTING_COMPLETE",RequestStatus.COMPLETED.value,audit_note)
    notify(req.requester.email,f"Đã có mã vật tư: {req.item_name}",req,f"Mã vật tư mới: {req.result_item_code}");return req
@router.post("/requests/{rid}/return",response_model=RequestOut)
def return_to_masterdata(rid:UUID,payload:ReasonIn,db:Session=Depends(get_db),user:User=Depends(allowed)):
    req=load(db,rid)
    if req.status!=RequestStatus.MASTERDATA_APPROVED.value:raise HTTPException(409,"Trạng thái không thể trả lại")
    req=transition(db,req,user,"ACCOUNTING_RETURN",RequestStatus.ACCOUNTING_RETURNED.value,payload.reason);notify(req.requester.email,f"Kế toán trả yêu cầu: {req.item_name}",req,f"Lý do: {payload.reason}");return req
