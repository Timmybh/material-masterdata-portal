import argparse,csv,json
from datetime import datetime,timedelta
from decimal import Decimal,InvalidOperation
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from unidecode import unidecode
from app.db import SessionLocal,init_db
from app.models import Item

COLUMNS={"Mã mới":"new_code","Mã cũ":"old_code","Id":"id","Code":"code","ParentId":"parent_id","IsGroup":"is_group","__unnamed_6":"source_extra_1","__unnamed_7":"source_extra_2","Name":"name","ItemTypeName":"item_type_name","Name2":"name2","ItemCustom":"item_custom","IsCustoms":"is_customs","IsItemWithColor":"is_item_with_color","IsItemWithSize":"is_item_with_size","IsItemWithArt":"is_item_with_art","IsItemWithProductCostId":"is_item_with_product_cost_id","IsItemWithBizDocId_C2":"is_item_with_biz_doc_id_c2","IsItemWithSymmetrical":"is_item_with_symmetrical","IsItemWithColorProduct":"is_item_with_color_product","ParentCode":"parent_code","ItemGroupCode":"item_group_code","KindCode":"kind_code","CustomerCode":"customer_code","ProductCostInfo":"product_cost_info","ProductItemCode":"product_item_code","BranchCode":"branch_code","IsMaterial":"is_material","UnitPrice":"unit_price","IsActive":"is_active","CreatedBy":"source_created_by","CreatedAt":"source_created_at","ModifiedBy":"source_modified_by","ModifiedAt":"source_modified_at","_SelectKey__cumontli":"source_select_key"}
BOOL_FIELDS={"is_group","is_customs","is_item_with_color","is_item_with_size","is_item_with_art","is_item_with_product_cost_id","is_item_with_biz_doc_id_c2","is_item_with_symmetrical","is_item_with_color_product","is_material","is_active","source_select_key"}
DATE_FIELDS={"source_created_at","source_modified_at"}; INT_FIELDS={"id","parent_id","source_created_by","source_modified_by"}
def clean(v):
    if v is None or str(v).strip().upper() in {"","NULL","NAN"}:return None
    return v.strip() if isinstance(v,str) else v
def as_bool(v):return str(v).strip().lower() in {"1","true","yes","x"}
def as_date(v):
    if v is None:return None
    if isinstance(v,datetime):return v
    if isinstance(v,(int,float)):return datetime(1899,12,30)+timedelta(days=float(v))
    try:return datetime.fromisoformat(str(v))
    except ValueError:return None
def norm(*parts):return " ".join(unidecode(str(x)).lower().strip() for x in parts if clean(x) is not None)
def source_rows(path):
    if Path(path).suffix.lower()==".csv":
        f=open(path,newline="",encoding="utf-8-sig")
        reader=csv.reader(f)
        return next(reader),reader,f
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True)
    return next(rows),rows,wb
def run(path:str):
    init_db();raw_headers,rows,source=source_rows(path)
    headers=[h if h not in (None,"") else f"__unnamed_{i}" for i,h in enumerate(raw_headers)];pos={h:i for i,h in enumerate(headers)}
    missing={"Id","Code","Name","ItemTypeName"}-set(pos)
    if missing:raise ValueError(f"Thiếu cột bắt buộc: {', '.join(sorted(missing))}")
    batch=[];total=skipped=0
    with SessionLocal() as db:
      for row in rows:
        data={dest:clean(row[pos[src]]) for src,dest in COLUMNS.items() if src in pos}
        if data.get("id") is None or data.get("code") is None or data.get("name") is None:skipped+=1;continue
        for k in BOOL_FIELDS:data[k]=as_bool(data.get(k))
        for k in DATE_FIELDS:data[k]=as_date(data.get(k))
        for k in INT_FIELDS:
            if data.get(k) is not None:data[k]=int(data[k])
        for k in ("source_extra_1","source_extra_2"):
            if data.get(k) is not None:data[k]=str(data[k])
        if data.get("unit_price") is not None:
            try:data["unit_price"]=Decimal(str(data["unit_price"]))
            except InvalidOperation:data["unit_price"]=None
        extras={str(h):row[i] for i,h in enumerate(headers) if h not in COLUMNS and h and clean(row[i]) is not None}
        data["extra_data"]=json.dumps(extras,ensure_ascii=False,default=str) if extras else None
        data["search_text"]=norm(*(data.get(k) for k in ("new_code","old_code","code","name","name2","item_custom","item_type_name","parent_code","item_group_code","kind_code","customer_code","product_cost_info","branch_code")))
        batch.append(data)
        if len(batch)>=1000:
            stmt=insert(Item).values(batch);db.execute(stmt.on_conflict_do_update(index_elements=[Item.id],set_={k:getattr(stmt.excluded,k) for k in batch[0] if k!="id"}));db.commit();total+=len(batch);print(f"Imported {total}");batch=[]
      if batch:
        stmt=insert(Item).values(batch);db.execute(stmt.on_conflict_do_update(index_elements=[Item.id],set_={k:getattr(stmt.excluded,k) for k in batch[0] if k!="id"}));db.commit();total+=len(batch)
    source.close();print(f"Done: {total} items; skipped: {skipped}")
if __name__=="__main__":
    p=argparse.ArgumentParser();p.add_argument("xlsx");run(p.parse_args().xlsx)
